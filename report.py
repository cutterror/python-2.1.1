from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import string
from statistic import Statistic


class Report:
    title_font = Font(name='Calibri', size=11, bold=True)

    border = Border(left=Side(border_style="thin", color='FF000000'),
                    right=Side(border_style="thin", color='FF000000'),
                    top=Side(border_style="thin", color='FF000000'),
                    bottom=Side(border_style="thin", color='FF000000'))

    def __init__(self, statistic: Statistic):
        self.__statistic = statistic
        self.__book = Workbook()
        self.__year_list = self.__book.active
        self.__year_list.title = "Статистика по годам"
        self.__city_list = self.__book.create_sheet("Статистика по городам")

    def generate_excel(self):
        self.print_columns(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.__statistic.selected_vacancy,
                            'Количество вакансий', 'Количество вакансий - ' + self.__statistic.selected_vacancy],
                           (list(self.__statistic.salary_dynamics.keys()), 'right', False),
                           (list(self.__statistic.salary_dynamics.values()), 'right', False),
                           (list(self.__statistic.selected_salary_dynamics.values()), 'right', False),
                           (list(self.__statistic.num_vacancies_dynamics.values()), 'right', False),
                           (list(self.__statistic.selected_num_vacancies_dynamics.values()), 'right', False))
        self.__book.active = self.__city_list
        self.print_columns(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'],
                           (list(self.__statistic.city_salary_dynamics.keys()), 'left', False),
                           (list(self.__statistic.city_salary_dynamics.values()), 'right', False),
                           (list(), 'right', False),
                           (list(self.__statistic.city_num_vacancies_dynamics.keys()), 'left', False),
                           (list(self.__statistic.city_num_vacancies_dynamics.values()), 'right', True))
        self.__book.active = self.__year_list
        self.__book.save("report.xlsx")

    def print_columns(self, titles: list, *args):
        self.fill_titles(titles)
        columns = string.ascii_uppercase[:len(titles)]
        for i, arg in enumerate(args):
            self.print_column(columns[i], arg[0], arg[1], arg[2])

    def print_column(self, column_name: str, column_data: list, alignment: str, is_percent: bool):
        ws = self.__book.active
        for i in range(len(column_data)):
            ws[column_name + str(i + 2)] = str(round(column_data[i] * 100, 2)) + '%' if is_percent else column_data[i]
            if ws.column_dimensions[column_name].width < len(str(column_data[i])) + 2:
                ws.column_dimensions[column_name].width = len(str(column_data[i])) + 2
            ws[column_name + str(i + 2)].alignment = Alignment(horizontal=alignment)
            ws[column_name + str(i + 2)].border = self.border

    def fill_titles(self, titles: list):
        columns = string.ascii_uppercase[:len(titles)]
        ws = self.__book.active
        for i, column in enumerate(columns):
            if titles[i] == "":
                ws.column_dimensions[column].width = 2
                continue
            ws[column + '1'] = titles[i]
            ws[column + '1'].font = self.title_font
            ws[column + '1'].border = self.border
            ws.column_dimensions[column].width = len(titles[i]) + 2
